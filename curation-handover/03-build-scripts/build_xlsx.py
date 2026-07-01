# -*- coding: utf-8 -*-
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC="/sessions/confident-intelligent-newton/mnt/outputs/filmcurio_candidates.csv"
rows=[]
with open(SRC,encoding="utf-8") as f:
    for r in csv.DictReader(f):
        rows.append(r)

# sort: Wave, Bucket(custom order), Year
border_order={"Canon backfill":0,"Contemporary curation":1,"Popular/genre depth":2}
rows.sort(key=lambda r:(int(r["Wave"]), border_order.get(r["Bucket"],9), r["Film_Title"]))

NAVY="1F2A44"; INK="222222"; GREY="6B7280"
F_CANON="E8EEF7"; F_CONTEMP="EAF3EC"; F_POP="FBF0E7"
BUCKET_FILL={"Canon backfill":F_CANON,"Contemporary curation":F_CONTEMP,"Popular/genre depth":F_POP}
HEAD_FILL=PatternFill("solid",fgColor=NAVY)
thin=Side(style="thin",color="D9DCE3")
BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
FONT="Arial"

wb=Workbook()

# ---------------- Sheet 1: Strategy ----------------
s=wb.active; s.title="전략 요약"
s.sheet_view.showGridLines=False
s.column_dimensions["A"].width=3
s.column_dimensions["B"].width=34
s.column_dimensions["C"].width=58
s.column_dimensions["D"].width=14

def put(cell,val,size=10,bold=False,color=INK,wrap=False,fill=None,align="left"):
    c=s[cell]; c.value=val
    c.font=Font(name=FONT,size=size,bold=bold,color=color)
    c.alignment=Alignment(horizontal=align,vertical="center",wrap_text=wrap)
    if fill: c.fill=PatternFill("solid",fgColor=fill)
    return c

s.merge_cells("B2:D2"); put("B2","FilmCurio / Metatake — 영화 카탈로그 확장 전략",16,True,NAVY)
s.merge_cells("B3:D3"); put("B3","작성일 2026-06-16 · 기준 데이터: metatake_films_567.csv + figures_takes_4662.csv",9,False,GREY)

r=5
def section(title):
    global r
    s.merge_cells(f"B{r}:D{r}")
    put(f"B{r}",title,11,True,"FFFFFF",fill=NAVY); r+=1
def line(label,val="",lbold=False):
    global r
    put(f"B{r}",label,10,lbold,INK,wrap=True)
    if val!="":
        s.merge_cells(f"C{r}:D{r}"); put(f"C{r}",val,10,False,INK,wrap=True)
    r+=1

section("현황")
line("현재 카탈로그","567편 (실질 ~555편 — 중복 8편 + TMDB 오매칭 노이즈 ~10편 정리 필요)")
line("최종 목표","1,000편 (MASTER.md 명시)")
line("이 리스트","신규 후보 405편 · 기존과 중복 0 · 405편 추가 시 ≈ 960편")
r+=1
section("진단 — 핵심 빈틈")
line("시대 편향","1990–2010년대에 439편 집중. 1980년 이전 거의 공백(1970s 6편, 1960s 1편) = 최대 구멍")
line("'상징 없는 유명작' 우려","데이터상 거의 없음 — 형상 산출량 평균 8.2개/편, 대중영화도 풍부(오션스11=17, 300=15)")
line("지역 빈틈","인도(Ray)·아프리카(Sembène)·대만/홍콩 뉴웨이브·고전 유럽 정전")
r+=1
section("큐레이션 기준 — 모든 후보가 동시 통과")
line("① 도달성","Criterion · Mubi · Kanopy · TSPDT 1000 · Sight&Sound · 영화제 수상 · 스트리밍 인기 중 ≥1곳 큐레이션")
line("② 밀도","학술 비평 문헌 존재 = figure-enrich 엔진의 원료 (유명 ≠ 충분, 비평이 써진 작품)")
line("③ 그래프 응집","기존 메타테이크·오퇴르에 연결되는 작품 우선 (고아 노드 회피) — Why 열에 표기")
r+=1
section("3 바구니 (가중치)")
line("클래식 정전 backfill","164편 (40%) — 유명 × 상징 밀도 최고 교집합. 이론의 원천 텍스트들")
line("동시대 큐레이션","132편 (33%) — 2018–2026 영화제·스트리밍·A24. 신선도·SEO. 카탈로그 오퇴르 심화")
line("대중·장르 깊이","109편 (27%) — 호러 정전·프랜차이즈·애니·다큐. 검색 트래픽 + 입증된 읽기 산출")
r+=1
section("3 웨이브 (각 ~135편, 내부 균형)")
line("Wave 1","가장 유명·연결성 높은 마키 타이틀 (즉시 트래픽·그래프 응집)")
line("Wave 2 / 3","점진 심화 — 깊은 정전·지역 뉴웨이브·롱테일")
line("각 웨이브 전 (런북 §4)","figure-enrich DRY 검수 → 스테이징 5–10편 → 전체 import. DB 백업·롤백 쿼리 준비")
r+=1
section("운영 순서")
line("1) 위생","중복 8 + 노이즈 ~10 제거 → run-tmdb-fetch-all (genres/overview 백필)")
line("2) 웨이브 파이프라인","import → figure-enrich → consolidate → author → rank → recommend")
line("3) 색인 게이트","take ≥3 → index · <3 → noindex(thin-content 회피)")
for rr in range(2, r+1): s.row_dimensions[rr].height=20

# ---------------- Sheet 2: Candidates ----------------
c2=wb.create_sheet("후보 405")
c2.sheet_view.showGridLines=False
heads=["Wave","바구니","제목 (Title)","감독 (Director)","연도","국가","선정 근거 (Why)"]
widths=[7,22,40,26,8,16,52]
for i,(h,w) in enumerate(zip(heads,widths),1):
    cell=c2.cell(row=1,column=i,value=h)
    cell.font=Font(name=FONT,size=10,bold=True,color="FFFFFF")
    cell.fill=HEAD_FILL; cell.alignment=Alignment(horizontal="center",vertical="center")
    cell.border=BORDER
    c2.column_dimensions[get_column_letter(i)].width=w
bucket_kr={"Canon backfill":"정전 backfill","Contemporary curation":"동시대 큐레이션","Popular/genre depth":"대중·장르"}
for ri,r0 in enumerate(rows,start=2):
    fill=BUCKET_FILL.get(r0["Bucket"],"FFFFFF")
    vals=[int(r0["Wave"]),bucket_kr.get(r0["Bucket"],r0["Bucket"]),r0["Film_Title"],
          r0["Film_Director_Name"],str(r0["Year"]),r0["Country"],r0["Why_curated"]]
    for ci,v in enumerate(vals,1):
        cell=c2.cell(row=ri,column=ci,value=v)
        cell.font=Font(name=FONT,size=10,color=INK)
        cell.fill=PatternFill("solid",fgColor=fill)
        cell.border=BORDER
        if ci in (1,5): cell.alignment=Alignment(horizontal="center",vertical="center")
        elif ci==7: cell.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)
        else: cell.alignment=Alignment(horizontal="left",vertical="center")
c2.freeze_panes="A2"
c2.auto_filter.ref=f"A1:G{len(rows)+1}"

# ---------------- Sheet 3: Wave plan (COUNTIFS) ----------------
w3=wb.create_sheet("웨이브 계획")
w3.sheet_view.showGridLines=False
for col,wd in zip("ABCDE",[16,18,20,16,12]): w3.column_dimensions[col].width=wd
w3.merge_cells("A1:E1")
t=w3["A1"]; t.value="웨이브 × 바구니 분포 (Candidates 시트 실시간 집계)"
t.font=Font(name=FONT,size=12,bold=True,color=NAVY)
hdr=["","정전 backfill","동시대 큐레이션","대중·장르","합계"]
for i,h in enumerate(hdr,1):
    cell=w3.cell(row=3,column=i,value=h)
    cell.font=Font(name=FONT,size=10,bold=True,color="FFFFFF"); cell.fill=HEAD_FILL
    cell.alignment=Alignment(horizontal="center"); cell.border=BORDER
rng_w="'후보 405'!$A$2:$A$"+str(len(rows)+1)
rng_b="'후보 405'!$B$2:$B$"+str(len(rows)+1)
bkr=["정전 backfill","동시대 큐레이션","대중·장르"]
for wi in range(1,4):
    rrow=3+wi
    c=w3.cell(row=rrow,column=1,value=f"Wave {wi}")
    c.font=Font(name=FONT,size=10,bold=True); c.border=BORDER
    for bi,bk in enumerate(bkr,2):
        col=get_column_letter(bi)
        cell=w3.cell(row=rrow,column=bi,value=f'=COUNTIFS({rng_w},{wi},{rng_b},"{bk}")')
        cell.alignment=Alignment(horizontal="center"); cell.border=BORDER; cell.font=Font(name=FONT,size=10)
    w3.cell(row=rrow,column=5,value=f"=SUM(B{rrow}:D{rrow})").border=BORDER
    w3.cell(row=rrow,column=5).alignment=Alignment(horizontal="center")
    w3.cell(row=rrow,column=5).font=Font(name=FONT,size=10,bold=True)
tot=7
w3.cell(row=tot,column=1,value="합계").font=Font(name=FONT,size=10,bold=True)
w3.cell(row=tot,column=1).border=BORDER
for bi in range(2,6):
    col=get_column_letter(bi)
    cell=w3.cell(row=tot,column=bi,value=f"=SUM({col}4:{col}6)")
    cell.alignment=Alignment(horizontal="center"); cell.font=Font(name=FONT,size=10,bold=True); cell.border=BORDER
    cell.fill=PatternFill("solid",fgColor="EEF1F6")

# ---------------- Sheet 4: Import (3-col, pipeline-ready) ----------------
im=wb.create_sheet("Import (3열)")
im.sheet_view.showGridLines=False
note=im["A1"]; im.merge_cells("A1:C1")
note.value="metatake_films_567.csv 와 동일 스키마. TMDB_ID는 importer가 제목+감독+연도로 해소. CSV: metatake_films_expansion_405.csv"
note.font=Font(name=FONT,size=9,color=GREY); note.alignment=Alignment(wrap_text=True)
im.row_dimensions[1].height=28
for i,h in enumerate(["Film_TMDB_ID","Film_Title","Film_Director_Name"],1):
    cell=im.cell(row=2,column=i,value=h)
    cell.font=Font(name=FONT,size=10,bold=True,color="FFFFFF"); cell.fill=HEAD_FILL
    cell.border=BORDER
for i,w in enumerate([16,40,28],1): im.column_dimensions[get_column_letter(i)].width=w
for ri,r0 in enumerate(rows,start=3):
    im.cell(row=ri,column=1,value="")
    im.cell(row=ri,column=2,value=r0["Film_Title"]).font=Font(name=FONT,size=10)
    im.cell(row=ri,column=3,value=r0["Film_Director_Name"]).font=Font(name=FONT,size=10)
im.freeze_panes="A3"

OUT="/sessions/confident-intelligent-newton/mnt/outputs/FilmCurio_확장후보_405.xlsx"
wb.save(OUT)

# import-ready CSV (3 columns, matching seed schema)
IMP="/sessions/confident-intelligent-newton/mnt/outputs/metatake_films_expansion_405.csv"
with open(IMP,"w",newline="",encoding="utf-8") as f:
    wcsv=csv.writer(f); wcsv.writerow(["Film_TMDB_ID","Film_Title","Film_Director_Name"])
    for r0 in rows: wcsv.writerow(["",r0["Film_Title"],r0["Film_Director_Name"]])
print("xlsx ->",OUT)
print("import csv ->",IMP)
print("rows:",len(rows))
