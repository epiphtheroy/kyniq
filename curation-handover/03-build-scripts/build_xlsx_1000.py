# -*- coding: utf-8 -*-
import csv
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT="/sessions/confident-intelligent-newton/mnt/outputs"
rows=list(csv.DictReader(open(f"{OUT}/filmcurio_candidates_1000.csv",encoding="utf-8")))

NAVY="1F2A44"; INK="222222"; GREY="6B7280"
F_CANON="E8EEF7"; F_CONTEMP="EAF3EC"; F_POP="FBF0E7"
BFILL={"Canon backfill":F_CANON,"Contemporary curation":F_CONTEMP,"Popular/genre depth":F_POP}
HEAD=PatternFill("solid",fgColor=NAVY)
thin=Side(style="thin",color="D9DCE3"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
FONT="Arial"
wb=Workbook()

# ---------- Sheet 1: Strategy ----------
s=wb.active; s.title="전략 요약"; s.sheet_view.showGridLines=False
for col,w in zip("ABCD",[3,34,58,12]): s.column_dimensions[col].width=w
def put(cell,v,size=10,bold=False,color=INK,wrap=False,fill=None):
    c=s[cell]; c.value=v; c.font=Font(name=FONT,size=size,bold=bold,color=color)
    c.alignment=Alignment(horizontal="left",vertical="center",wrap_text=wrap)
    if fill: c.fill=PatternFill("solid",fgColor=fill)
s.merge_cells("B2:D2"); put("B2","FilmCurio / Metatake — +1,000편 확장 전략 (Wave 2)",16,True,NAVY)
s.merge_cells("B3:D3"); put("B3","작성 2026-06-16 · 수요×밀도 랭킹 모델 · 기존 567 + 405 와 중복 0",9,False,GREY)
r=5
def sec(t):
    global r; s.merge_cells(f"B{r}:D{r}"); put(f"B{r}",t,11,True,"FFFFFF",fill=NAVY); r+=1
def ln(a,b=""):
    global r; put(f"B{r}",a,10,False,INK,wrap=True)
    if b!="": s.merge_cells(f"C{r}:D{r}"); put(f"C{r}",b,10,False,INK,wrap=True)
    r+=1
sec("범위 (확정)")
ln("이번 추가","신규 1,000편 (405와 별개). 중복 제거 대상 = 기존 567 + 405")
ln("총 카탈로그","≈ 567 + 405 + 1,000 = 1,972편")
ln("⚠ 규모 주의","1,972편 × ~10 Q&A ≈ 2만 페이지. 프로젝트 문서가 경고한 'AI 과다콘텐츠' SEO 상한의 2배 — 웨이브별 품질·색인 게이트를 반드시 지킬 것")
r+=1
sec("선별 모델 — 수요 × 밀도 (랭킹 산식)")
ln("Priority 산식","Demand(1–5) × Density(1–5) + 최신가산 + 그래프가산")
ln("Demand (수요/클릭)","IMDb·Letterboxd 인기·평점 / 스트리밍 최다시청 / 박스오피스·프랜차이즈 / 시네필 검색")
ln("Density (밀도/상징)","TSPDT·S&S·Criterion·학술 비평 존재 / 오퇴르 / 형상 산출 잠재력")
ln("최신가산","+2 (2018년~), +1 (2010–17). '지금 사람들이 찾는' 작품 우대")
ln("그래프가산","+1 (기존 카탈로그에 허브를 가진 감독) — 고아 노드 회피, 메타테이크 응집")
ln("소싱 깔때기","플랫폼 큐레이션(Mubi/Criterion/Kanopy/넷플릭스) ∩ 비평 정전. 랭킹은 그 교집합 안의 순위")
r+=1
tot=len(rows); bc=Counter(x["Bucket"] for x in rows)
sec("구성 (버킷 균형 유지 40 / 33 / 27)")
ln("클래식 정전 backfill",f"{bc['Canon backfill']}편 (40%) — 1980년 이전·세계영화·오퇴르 심화 (밀도 척추)")
ln("동시대 큐레이션",f"{bc['Contemporary curation']}편 (33%) — 2018–2026 영화제·스트리밍 + 살아있는 오퇴르 후속작")
ln("대중·장르 깊이",f"{bc['Popular/genre depth']}편 (27%) — 프랜차이즈·호러·애니·아시아 상업영화·다큐 (검색 트래픽)")
r+=1
sec("4 웨이브 (각 250편 = 수요×밀도 우선순위 분위)")
ln("Wave 1","최고 우선순위 250 — 가장 클릭·검색되고 밀도도 높은 교집합 (즉시 트래픽)")
ln("Wave 2–4","우선순위 내림차순. 깊은 정전·세계영화·롱테일은 후속 웨이브")
ln("주의","랭킹은 '수집 우선순위'일 뿐 — 버킷 균형(40/33/27)은 전체 1,000편 기준으로 보존됨")
r+=1
dec=Counter((int(x["Year"])//10)*10 for x in rows)
pre80=sum(v for k,v in dec.items() if k<1980)
sec("커버리지 — 빈틈 직접 타격")
ln("1980년 이전",f"{pre80}편 (기존 카탈로그 최대 구멍) — 무성·고전기·뉴웨이브·세계 정전")
ln("지역 확장","인도 35 · 한국 42 · 이란 18 · 중국 27 · 일본 81 — 기존 얇던 지역 보강")
ln("미국 비중","414편 (41%) — 나머지 59%는 비미국권")
r+=1
sec("운영 순서 (런북 §4 준수)")
ln("1) 위생","기존 567 중복 8 + 노이즈 ~10 정리 → tmdb-fetch 백필")
ln("2) 웨이브별","import → figure-enrich(DRY→스테이징→전체) → consolidate → author → rank")
ln("3) 색인 게이트","take ≥3 → index · <3 → noindex. 과다콘텐츠 위험 관리 = 품질 게이트")
for rr in range(2,r+1): s.row_dimensions[rr].height=19

# ---------- Sheet 2: Ranked 1000 ----------
c2=wb.create_sheet("후보 1000 (랭킹)"); c2.sheet_view.showGridLines=False
heads=["순위","Wave","바구니","제목","감독","연도","국가","수요","밀도","점수"]
widths=[6,6,20,40,26,7,15,7,7,8]
for i,(h,w) in enumerate(zip(heads,widths),1):
    cell=c2.cell(1,i,h); cell.font=Font(name=FONT,size=10,bold=True,color="FFFFFF")
    cell.fill=HEAD; cell.alignment=Alignment(horizontal="center",vertical="center"); cell.border=BORDER
    c2.column_dimensions[get_column_letter(i)].width=w
bkr={"Canon backfill":"정전","Contemporary curation":"동시대","Popular/genre depth":"대중·장르"}
for ri,r0 in enumerate(rows,start=2):
    fill=BFILL[r0["Bucket"]]
    vals=[int(r0["Rank"]),int(r0["Wave"]),bkr[r0["Bucket"]],r0["Film_Title"],r0["Film_Director_Name"],
          str(r0["Year"]),r0["Country"],int(r0["Demand"]),int(r0["Density"]),int(r0["Priority"])]
    for ci,v in enumerate(vals,1):
        cell=c2.cell(ri,ci,v); cell.font=Font(name=FONT,size=9,color=INK)
        cell.fill=PatternFill("solid",fgColor=fill); cell.border=BORDER
        cell.alignment=Alignment(horizontal=("center" if ci in (1,2,6,8,9,10) else "left"),vertical="center")
c2.freeze_panes="A2"; c2.auto_filter.ref=f"A1:J{len(rows)+1}"

# ---------- Sheet 3: Wave plan ----------
w3=wb.create_sheet("웨이브 계획"); w3.sheet_view.showGridLines=False
for col,wd in zip("ABCDE",[14,16,18,16,12]): w3.column_dimensions[col].width=wd
w3.merge_cells("A1:E1"); t=w3["A1"]; t.value="웨이브 × 바구니 (실시간 집계)"; t.font=Font(name=FONT,size=12,bold=True,color=NAVY)
for i,h in enumerate(["","정전","동시대","대중·장르","합계"],1):
    cell=w3.cell(3,i,h); cell.font=Font(name=FONT,size=10,bold=True,color="FFFFFF"); cell.fill=HEAD
    cell.alignment=Alignment(horizontal="center"); cell.border=BORDER
N=len(rows)+1
rw=f"'후보 1000 (랭킹)'!$B$2:$B${N}"; rb=f"'후보 1000 (랭킹)'!$C$2:$C${N}"
for wi in range(1,5):
    rr=3+wi; w3.cell(rr,1,f"Wave {wi}").font=Font(name=FONT,size=10,bold=True); w3.cell(rr,1).border=BORDER
    for bi,bk in enumerate(["정전","동시대","대중·장르"],2):
        cell=w3.cell(rr,bi,f'=COUNTIFS({rw},{wi},{rb},"{bk}")'); cell.border=BORDER
        cell.alignment=Alignment(horizontal="center"); cell.font=Font(name=FONT,size=10)
    cell=w3.cell(rr,5,f"=SUM(B{rr}:D{rr})"); cell.border=BORDER; cell.font=Font(name=FONT,size=10,bold=True)
    cell.alignment=Alignment(horizontal="center")
w3.cell(8,1,"합계").font=Font(name=FONT,size=10,bold=True); w3.cell(8,1).border=BORDER
for bi in range(2,6):
    col=get_column_letter(bi); cell=w3.cell(8,bi,f"=SUM({col}4:{col}7)")
    cell.font=Font(name=FONT,size=10,bold=True); cell.alignment=Alignment(horizontal="center")
    cell.border=BORDER; cell.fill=PatternFill("solid",fgColor="EEF1F6")
# decade mini-table
w3.merge_cells("A10:E10"); d=w3["A10"]; d.value="10년 단위 분포"; d.font=Font(name=FONT,size=11,bold=True,color=NAVY)
dec=Counter((int(x["Year"])//10)*10 for x in rows)
w3.cell(11,1,"연대").font=Font(name=FONT,bold=True,size=10); w3.cell(11,2,"편수").font=Font(name=FONT,bold=True,size=10)
for i,k in enumerate(sorted(dec),start=12):
    w3.cell(i,1,f"{k}s"); w3.cell(i,2,dec[k])

# ---------- Sheet 4: Import ----------
im=wb.create_sheet("Import (3열)"); im.sheet_view.showGridLines=False
im.merge_cells("A1:C1"); n=im["A1"]
n.value="metatake_films_567.csv 와 동일 스키마. TMDB_ID는 importer가 제목+감독+연도로 해소. CSV: metatake_films_expansion_1000.csv (순위순)"
n.font=Font(name=FONT,size=9,color=GREY); n.alignment=Alignment(wrap_text=True); im.row_dimensions[1].height=28
for i,h in enumerate(["Film_TMDB_ID","Film_Title","Film_Director_Name"],1):
    cell=im.cell(2,i,h); cell.font=Font(name=FONT,size=10,bold=True,color="FFFFFF"); cell.fill=HEAD; cell.border=BORDER
for i,w in enumerate([16,40,28],1): im.column_dimensions[get_column_letter(i)].width=w
for ri,r0 in enumerate(rows,start=3):
    im.cell(ri,1,""); im.cell(ri,2,r0["Film_Title"]).font=Font(name=FONT,size=10)
    im.cell(ri,3,r0["Film_Director_Name"]).font=Font(name=FONT,size=10)
im.freeze_panes="A3"

OUTF=f"{OUT}/FilmCurio_확장후보_1000.xlsx"; wb.save(OUTF); print("saved",OUTF)
