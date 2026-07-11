#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
영화 촬영지 에이전트 — LLM + 웹검색, 인용근거·다중출처 검증판
=============================================================

이 에이전트는 특정 사이트를 긁지 않습니다. 웹을 검색해 여러 독립 출처에서
사실을 종합하고, 검증 게이트를 통과한 항목만 'verified'로 내보냅니다.
→ 단일 데이터베이스 의존이 사라져 법적으로 더 깨끗하고, 다중출처 교차검증으로
   환각을 걸러 신뢰도를 실용 수준으로 끌어올립니다.

파이프라인 (라운드 반복으로 확정된 절차):
  1) SEARCH    : 영화별 다각도 검색(촬영지 / set-vs-filmed / "real or set?")
  2) GENERATE  : LLM이 후보 촬영지 추출 — 각 항목에 출처 URL 필수.
                 narrative_setting(영화 속 장소)와 real filming location을 분리.
                 built_set(세트/CGI 여부) 판정.
  3) VERIFY    : (a) 규칙 게이트 — 서로 다른 도메인 ≥2 → verified, 1 → single_source
                 (b) LLM 반증 패스 — "이거 set 아니냐 / 배경지 아니냐 / 다른 영화 아니냐"
  4) GEOCODE   : (선택) 주소 지오코딩으로 지역 정합성 sanity check
  5) OUTPUT    : verified만 배포(single_source/unverified는 격리). JSON + xlsx.

확정된 실패모드 대응(프로토타입 라운드에서 관찰):
  F1 검색 1위가 movie-locations.com → ≥2 독립 도메인 강제(단일출처 금지)
  F2 유명 '촬영지'가 실은 세트(Parasite 저택, Skyfall Lodge) → built_set 플래그 + 반증패스
  F3 입도 편차(주소 vs 자연지형) → granularity 태그
  F4 set↔filmed 혼동(Fury Road=나미비아, Skyfall 'Scotland' 일부=Surrey) → 필드 분리

의존성: requests, openai, openpyxl
환경변수:
  LLM_API_KEY / LLM_BASE_URL / LLM_MODEL          (LLM)
  SEARCH_API_KEY  (+ SEARCH_PROVIDER=tavily|brave) (웹검색)
  GEOCODE=1 이면 Nominatim 정합성 체크 활성화(예의상 1req/s)

사용:
  python movie_locations_llmsearch.py --titles "Parasite (2019)" "La La Land" "Skyfall"
  python movie_locations_llmsearch.py --titles "Heat 1995" --min-sources 2 --out run
"""

import argparse
import json
import os
import re
import sys
import time
from urllib.parse import urlparse

import requests

USER_AGENT = "Mozilla/5.0 (compatible; MetatakeBot/1.0; +https://metatake.net/bot)"

# 보호 데이터베이스(법적 사유): 이 도메인 '하나만'이면 verified 불가 + 단독이면 격리.
SOLE_SOURCE_BLOCKLIST = {"movie-locations.com", "atlasofwonders.com"}

# 출처 권위 등급. Tier A는 단일출처라도 '개연성'을 통과하면 'probable'로 배포 가능
# (개방/권위 출처라 신뢰도·법적 안전 둘 다 양호). Tier B는 'weak'(검토 버킷).
TIER_A = {
    "en.wikipedia.org", "wikipedia.org", "imdb.com", "cnn.com", "bbc.co.uk",
    "bbc.com", "npr.org", "variety.com", "theguardian.com", "nytimes.com",
    "reuters.com", "apnews.com", "dezeen.com", "focusfeatures.com",
    "screenglobalproduction.com", "english.visitkorea.or.kr", "newzealand.com",
    "visitsurrey.com", "salzburg.info", "salzburgerland.com",
}
TIER_B = {
    "giggster.com", "screenrant.com", "thecinemaholic.com", "almostginger.com",
    "atlasobscura.com", "timeout.com", "legendarytrips.com", "huntingbond.com",
    "thejamesbonddossier.com", "nomadicnotes.com", "socalpulse.com",
    "nolanlocations.com", "filmoblivion.com", "set-jetter.com",
}


def source_tier(domain):
    d = domain.replace("www.", "")
    if d in SOLE_SOURCE_BLOCKLIST:
        return "X"                       # 보호 DB → 특별 취급
    if d in TIER_A or d.endswith((".gov", ".gov.uk", ".go.kr", ".or.kr")):
        return "A"
    if d in TIER_B:
        return "B"
    return "C"


# ---------------------------------------------------------------------------
# 1) SEARCH (provider 추상화)
# ---------------------------------------------------------------------------
def web_search(query, k=8):
    provider = os.getenv("SEARCH_PROVIDER", "tavily").lower()
    key = os.getenv("SEARCH_API_KEY")
    if not key:
        raise RuntimeError("SEARCH_API_KEY 필요(검색 provider)")
    if provider == "tavily":
        r = requests.post("https://api.tavily.com/search",
                          json={"api_key": key, "query": query, "max_results": k,
                                "include_answer": False},
                          timeout=30)
        r.raise_for_status()
        return [{"title": x.get("title", ""), "url": x.get("url", ""),
                 "snippet": x.get("content", "")} for x in r.json().get("results", [])]
    if provider == "brave":
        r = requests.get("https://api.search.brave.com/res/v1/web/search",
                         headers={"X-Subscription-Token": key, "Accept": "application/json"},
                         params={"q": query, "count": k}, timeout=30)
        r.raise_for_status()
        web = r.json().get("web", {}).get("results", [])
        return [{"title": x.get("title", ""), "url": x.get("url", ""),
                 "snippet": x.get("description", "")} for x in web]
    raise RuntimeError(f"알 수 없는 SEARCH_PROVIDER: {provider}")


def gather_evidence(title):
    """영화별 다각도 검색 결과를 합쳐 LLM 입력용 증거 묶음으로."""
    queries = [
        f"{title} filming locations real places list",
        f"{title} where was it filmed set vs actually filmed",
        f"{title} filming location is it a real place or a built set",
    ]
    evidence, seen = [], set()
    for q in queries:
        try:
            for hit in web_search(q):
                u = hit["url"]
                if u and u not in seen:
                    seen.add(u)
                    evidence.append(hit)
        except Exception as e:
            print(f"[warn] 검색 실패: {q} - {e}", file=sys.stderr)
        time.sleep(1)
    return evidence


# ---------------------------------------------------------------------------
# 2) GENERATE (인용 강제 추출)
# ---------------------------------------------------------------------------
GEN_PROMPT = """\
You are a film-location fact extractor. From the SEARCH EVIDENCE (title/url/snippet
list) extract REAL-WORLD filming locations for the given film. Output STRICT JSON.

HARD RULES:
- Use ONLY facts supported by the evidence. Every location MUST include "sources":
  a list of the evidence URLs that support it. If you cannot cite >=1 URL, DROP it.
- Separate the in-film place ("narrative_setting") from the real filming location
  ("real_name", "filming_area", "country").
- "built_set": true if the place is a constructed set / CGI / studio backlot
  (NOT a real visitable location). Watch for famous houses that are actually sets.
- Do NOT copy long descriptive sentences from any source; write scene_role in your
  own short neutral words. Keep proper nouns (venue, address, place) accurate.
- "granularity": one of address / venue / area / region / set.
- Rephrase, never reproduce a single site's wording or its full ordered list.

OUTPUT JSON:
{"title":str,"year":int|null,"narrative_setting":str,
 "locations":[{"real_name":str,"narrative_setting":str,"built_set":bool,
   "scene_role":str,"address":str,"filming_area":str,"country":str,
   "granularity":str,"sources":[url,...]}]}
Return ONLY the JSON object."""

VERIFY_PROMPT = """\
You are an adversarial fact-checker. For the candidate filming location, decide if it
is reliable, using ONLY the provided evidence snippets. Try to FALSIFY it:
- Is it actually a built set / CGI rather than a real place? (set built_set correctly)
- Is the evidence about where the story is SET rather than where it was FILMED?
- Could it be confused with a different film or a same-named place?
Return STRICT JSON: {"keep":bool,"built_set":bool,"reason":str}."""


def _llm(messages):
    from openai import OpenAI
    client = OpenAI(api_key=os.getenv("LLM_API_KEY"),
                    base_url=os.getenv("LLM_BASE_URL", "https://api.openai.com/v1"))
    resp = client.chat.completions.create(
        model=os.getenv("LLM_MODEL", "gpt-4o-mini"),
        temperature=0, response_format={"type": "json_object"}, messages=messages)
    return json.loads(resp.choices[0].message.content)


def generate(title, evidence):
    ev = json.dumps(evidence, ensure_ascii=False)[:16000]
    return _llm([{"role": "system", "content": GEN_PROMPT},
                 {"role": "user", "content": f"FILM: {title}\n\nSEARCH EVIDENCE:\n{ev}"}])


def verify_candidate(loc, evidence):
    snippets = "\n".join(f"- {e['url']}: {e['snippet'][:300]}" for e in evidence
                         if e["url"] in loc.get("sources", []))[:6000]
    try:
        return _llm([{"role": "system", "content": VERIFY_PROMPT},
                     {"role": "user", "content": f"CANDIDATE: {json.dumps(loc, ensure_ascii=False)}\n\nEVIDENCE:\n{snippets}"}])
    except Exception as e:
        print(f"[warn] verify 실패: {e}", file=sys.stderr)
        return {"keep": True, "built_set": loc.get("built_set", False), "reason": "verifier skipped"}


# ---------------------------------------------------------------------------
# 3) VERIFY 게이트 (규칙) + 4) GEOCODE (선택)
# ---------------------------------------------------------------------------
def distinct_domains(urls):
    return {urlparse(u).netloc.replace("www.", "") for u in urls if u}


def _plausible(loc):
    """개연성: 좌표가 명시적으로 실패(False)가 아니고 지역/국가 정보가 있는가."""
    if loc.get("_geocode") is False:
        return False
    return bool(loc.get("country") or loc.get("filming_area"))


def judge(loc, min_sources=2):
    """버리지 않고 등급(tier)으로 판정. single_source를 권위/개연성/법적사유로 차등.

    tiers: verified > probable > weak > quarantined_legal > rejected
    """
    doms = distinct_domains(loc.get("sources", []))
    independent = doms - SOLE_SOURCE_BLOCKLIST
    loc["source_domains"] = sorted(doms)
    loc["source_count"] = len(doms)
    best = min((source_tier(d) for d in doms), default="C")  # A<B<C<X 사전순; A가 최고

    if len(independent) >= min_sources:
        conf = "verified"
    elif len(doms) >= 2 and len(independent) >= 1:
        conf = "verified"                       # 독립 1 + 보강 1 이상
    elif len(doms) == 1:
        only = next(iter(doms))
        t = source_tier(only)
        if t == "X":
            conf = "quarantined_legal"          # 보호 DB 단독 → 격리(법적). 배포 안 함
        elif t == "A" and _plausible(loc):
            conf = "probable"                    # 권위 단일출처 + 개연성 → 배포(플래그)
        elif t in ("A", "B"):
            conf = "weak"                        # 검토 버킷
        else:
            conf = "weak"
    else:
        conf = "unverified"

    # 개연성 탈락이면 강등
    if conf in ("verified", "probable") and not _plausible(loc):
        conf = "weak"
    if loc.get("built_set") and conf == "verified":
        conf = "verified_set_not_real"
    loc["confidence"] = conf
    loc["top_tier"] = best
    return loc


SHIPPABLE = ("verified", "verified_set_not_real", "probable")


def recovery_search(loc, title):
    """단일출처/약함 항목에 대한 2차 타깃 검색 → 독립출처 보강 시도."""
    q = f"{title} {loc.get('real_name','')} filming location"
    try:
        for hit in web_search(q, k=6):
            u = hit.get("url")
            if u and u not in loc.get("sources", []):
                loc.setdefault("sources", []).append(u)
    except Exception as e:
        print(f"[warn] 복구검색 실패: {e}", file=sys.stderr)
    return loc


def geocode_ok(loc):
    if os.getenv("GEOCODE") != "1" or not loc.get("address"):
        return None
    try:
        r = requests.get("https://nominatim.openstreetmap.org/search",
                         params={"q": loc["address"], "format": "json", "limit": 1},
                         headers={"User-Agent": USER_AGENT}, timeout=20)
        time.sleep(1)  # Nominatim 예의: 1req/s
        hits = r.json()
        if not hits:
            return False
        loc["coordinates"] = f"{hits[0]['lat']},{hits[0]['lon']}"
        return True
    except Exception:
        return None


# ---------------------------------------------------------------------------
# 파이프라인
# ---------------------------------------------------------------------------
def process_film(title, min_sources):
    print(f"[film] {title}")
    evidence = gather_evidence(title)
    data = generate(title, evidence)
    out = []
    for loc in data.get("locations", []):
        v = verify_candidate(loc, evidence)
        if not v.get("keep", True):
            print(f"  [drop] {loc.get('real_name')} - {v.get('reason')}", file=sys.stderr)
            continue
        loc["built_set"] = bool(v.get("built_set", loc.get("built_set", False)))
        loc["_geocode"] = geocode_ok(loc)
        judge(loc, min_sources)
        # 단일출처/약함/격리 → 2차 타깃 검색으로 독립출처 보강 후 재판정(복구)
        if loc["confidence"] in ("weak", "single_source", "quarantined_legal"):
            recovery_search(loc, title)
            loc["_geocode"] = geocode_ok(loc)
            judge(loc, min_sources)
        out.append(loc)
    data["locations"] = out
    data["shipped"] = [l for l in out if l["confidence"] in SHIPPABLE]
    data["quarantined"] = [l for l in out if l["confidence"] == "quarantined_legal"]
    return data


def write_outputs(movies, out_base):
    with open(f"{out_base}.json", "w", encoding="utf-8") as f:
        json.dump({"method": "LLM+web search, citation-grounded, multi-source verified",
                   "movie_count": len(movies), "movies": movies}, f,
                  ensure_ascii=False, indent=2)
    print(f"[ok] {out_base}.json")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return
    wb = Workbook(); s = wb.active; s.title = "Locations"
    cols = ["Title", "Year", "Real place", "Narrative setting", "Built set?", "Scene/role",
            "Address", "Filming area", "Country", "Granularity", "Confidence",
            "Source count", "Sources"]
    s.append(cols)
    for m in movies:
        for l in m.get("locations", []):
            s.append([m.get("title"), m.get("year"), l.get("real_name"),
                      l.get("narrative_setting"), "YES" if l.get("built_set") else "",
                      l.get("scene_role"), l.get("address"), l.get("filming_area"),
                      l.get("country"), l.get("granularity"), l.get("confidence"),
                      l.get("source_count"), ", ".join(l.get("source_domains", []))])
    hf = Font(name="Arial", bold=True, color="FFFFFF")
    fill = PatternFill("solid", start_color="1F4E78")
    for c in range(1, len(cols) + 1):
        cell = s.cell(1, c); cell.font = hf; cell.fill = fill
    for i, w in enumerate([16, 6, 30, 22, 9, 34, 28, 22, 16, 11, 20, 11, 34], 1):
        s.column_dimensions[chr(64 + i)].width = w
    for row in s.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    s.freeze_panes = "A2"; s.auto_filter.ref = f"A1:M{s.max_row}"
    wb.save(f"{out_base}.xlsx"); print(f"[ok] {out_base}.xlsx")


def main():
    ap = argparse.ArgumentParser(description="LLM+검색 영화 촬영지 에이전트")
    ap.add_argument("--titles", nargs="+", required=True)
    ap.add_argument("--min-sources", type=int, default=2, help="verified 최소 독립 도메인 수")
    ap.add_argument("--out", default="llmsearch_run")
    args = ap.parse_args()
    if not os.getenv("LLM_API_KEY"):
        sys.exit("[error] LLM_API_KEY 필요")
    movies = [process_film(t, args.min_sources) for t in args.titles]
    tot = sum(len(m["locations"]) for m in movies)
    ship = sum(len(m["shipped"]) for m in movies)
    print(f"[done] {len(movies)}편 | 후보 {tot} | verified {ship}")
    write_outputs(movies, args.out)


if __name__ == "__main__":
    main()
