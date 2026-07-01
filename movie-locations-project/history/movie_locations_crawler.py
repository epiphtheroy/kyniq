#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
영화 촬영지 사실(fact) 수집 크롤러 — 멀티 소스 (LLM 정규화 + Wikidata)
=====================================================================

두 개의 소스를 지원하고, 병합(union + 중복제거)해서 자유롭게 쓸 수 있는
데이터셋을 만듭니다.

  소스 A: movie-locations.com  (--source movieloc)
     - 페이지별 저렴한 LLM이 사실만 추출·정규화.
     - scene_role 등 설명은 100% 리프레이즈(원문 표현 미사용).
     - 풍부한 fictional_name / scene_role 제공.
     - 주의: 데이터베이스권 리스크가 있어 '교차검증/일부' 용도로 권장.

  소스 B: Wikidata SPARQL (P915 filming location)  (--source wikidata)
     - 라이선스 CC0 (퍼블릭 도메인) → 재사용·상업화 자유.
     - 구조화 데이터라 LLM 불필요. 좌표(P625)·행정구역(P131)·국가(P17)·
       주소(P6375)까지 그대로 매핑.
     - fictional_name/scene_role은 대개 없음(빈 값) — 사실 그대로.

  병합: --source both → 영화 제목 기준으로 합치고, 촬영지는 real_name으로
        중복제거. movie-locations의 풍부한 설명을 유지하되 Wikidata의 좌표·
        주소로 빈칸을 채웁니다. 각 항목에 출처(sources)를 기록합니다.

공통 스키마(촬영지):
  fictional_name, real_name, scene_role, address, area, status, coordinates, sources

합법성 메모:
  - 크롤 가능 여부(robots.txt)와 재사용 가능 여부(저작권/DB권)는 별개입니다.
  - Wikidata(CC0)는 재사용 제약이 사실상 없습니다. 베이스로 권장.
  - movie-locations.com은 표현을 베끼지 않고(리프레이즈) 소량·교차검증 위주로.

의존성: requests, beautifulsoup4, openpyxl, openai
설정(LLM, movieloc 소스에만 필요):
  export LLM_API_KEY=...   LLM_BASE_URL=https://api.openai.com/v1   LLM_MODEL=gpt-4o-mini

사용 예:
  python movie_locations_crawler.py --source wikidata --titles "Heat" "Skyfall" "Inception"
  python movie_locations_crawler.py --source movieloc --letter h --limit 10
  python movie_locations_crawler.py --source both --titles "Heat" --urls https://movie-locations.com/movies/h/Heat.php
"""

import argparse
import json
import os
import re
import sys
import time
import urllib.parse
import urllib.robotparser as robotparser

import requests
from bs4 import BeautifulSoup

BASE = "https://movie-locations.com"
USER_AGENT = "MovieLocationsFactBot/3.0 (research; contact: your-email@example.com)"

WD_SPARQL = "https://query.wikidata.org/sparql"
WD_API = "https://www.wikidata.org/w/api.php"


def _norm(s):
    return re.sub(r"[^a-z0-9]+", "", (s or "").lower())


def _clean(t):
    return re.sub(r"\s+", " ", (t or "")).strip()


# ============================================================================
# 소스 A: movie-locations.com
# ============================================================================
class PoliteFetcher:
    def __init__(self, delay=2.0):
        self.delay = delay
        self.session = requests.Session()
        self.session.headers.update({"User-Agent": USER_AGENT})
        self.rp = robotparser.RobotFileParser()
        self.rp.set_url(f"{BASE}/robots.txt")
        try:
            self.rp.read()
        except Exception:
            print("[warn] robots.txt 읽기 실패 — 신중히 진행", file=sys.stderr)

    def allowed(self, url):
        try:
            return self.rp.can_fetch(USER_AGENT, url)
        except Exception:
            return True

    def get(self, url, retries=1):
        if not self.allowed(url):
            print(f"[skip] robots 비허용: {url}", file=sys.stderr)
            return None
        for attempt in range(retries + 1):
            try:
                r = self.session.get(url, timeout=20)
                if r.status_code == 200:
                    time.sleep(self.delay)
                    return r.text
                print(f"[warn] HTTP {r.status_code}: {url}", file=sys.stderr)
            except requests.RequestException as e:
                print(f"[warn] 요청 실패({attempt}): {url} - {e}", file=sys.stderr)
            time.sleep(self.delay)
        return None


def clean_page_text(html):
    soup = BeautifulSoup(html, "html.parser")
    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()
    parts = []
    h1 = soup.find("h1")
    if h1:
        parts.append("H1: " + _clean(h1.get_text()))
    for el in soup.find_all(["li", "p", "figcaption"]):
        t = _clean(el.get_text())
        if not t or re.fullmatch(r"[A-Z0-9](\s*[A-Z0-9])*", t):
            continue
        if t in ("HOME", "FILMS", "PLACES", "PEOPLE", "MORE", "SEARCH"):
            continue
        parts.append(t)
    return "\n".join(dict.fromkeys(parts))[:14000]


SYSTEM_PROMPT = """\
You extract ONLY objective, factual data about a film and its real-world shooting
locations from the page text. Output STRICT JSON. Rules:

ALLOWED FACTS ONLY: title, year, directors, cast, regions, and per location:
fictional_name, real_name, scene_role, address, area, status.

REPHRASING RULE (MANDATORY, 100%):
- NEVER copy the source's sentences or descriptive wording.
- Write scene_role in YOUR OWN concise neutral phrasing.
- Do NOT reproduce the author's commentary, jokes, opinions or trivia.
- EXCEPTION: proper nouns that ARE the fact (real venue names, place names,
  street addresses, person names, years) must stay ACCURATE and unaltered.

FIELDS: fictional_name = in-film name if any (else ""); real_name = actual venue;
scene_role = one short rephrased clause; address = street address or "";
area = neighbourhood/city/region; status = active/closed/demolished/rebuilt/
repurposed/renamed or "".

OUTPUT JSON: {"title":str,"year":int|null,"directors":[str],"cast":[str],
"regions":[str],"locations":[{"fictional_name":str,"real_name":str,
"scene_role":str,"address":str,"area":str,"status":str}]}
Return ONLY the JSON object."""


def llm_extract(page_text, url, model, api_key, base_url):
    from openai import OpenAI
    client = OpenAI(api_key=api_key, base_url=base_url)
    resp = client.chat.completions.create(
        model=model, temperature=0,
        response_format={"type": "json_object"},
        messages=[{"role": "system", "content": SYSTEM_PROMPT},
                  {"role": "user", "content": f"PAGE URL: {url}\n\nPAGE TEXT:\n{page_text}"}],
    )
    data = json.loads(resp.choices[0].message.content)
    data["source_url"] = url
    data["sources"] = ["movieloc"]
    for loc in data.setdefault("locations", []):
        for k in ("fictional_name", "real_name", "scene_role", "address", "area", "status"):
            loc.setdefault(k, "")
        loc["coordinates"] = ""
        loc["sources"] = ["movieloc"]
    return data


def movieloc_discover(fetcher, letter, limit=20):
    html = fetcher.get(f"{BASE}/movies/{letter}/{letter}-movies.php")
    if not html:
        return []
    soup = BeautifulSoup(html, "html.parser")
    urls = []
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if re.search(rf"/movies/{letter}/[^/]+\.php$", href) and "movies.php" not in href:
            full = urllib.parse.urljoin(f"{BASE}/movies/{letter}/", href)
            if full not in urls and not full.endswith(f"{letter}-movies.php"):
                urls.append(full)
        if len(urls) >= limit:
            break
    return urls


# ============================================================================
# 소스 B: Wikidata SPARQL (P915, CC0)
# ============================================================================
# 사용 속성: P915 filming location · P577 publication date · P57 director
#           P161 cast member · P131 admin area · P17 country · P625 coordinate
#           P6375 street address
WD_HEADERS = {"User-Agent": USER_AGENT, "Accept": "application/sparql-results+json"}


def wd_resolve_qid(title, session):
    """제목 → 영화 QID (wbsearchentities). 첫 결과 사용."""
    params = {"action": "wbsearchentities", "search": title, "language": "en",
              "type": "item", "format": "json", "limit": 5}
    r = session.get(WD_API, params=params, headers={"User-Agent": USER_AGENT}, timeout=20)
    r.raise_for_status()
    hits = r.json().get("search", [])
    return (hits[0]["id"], hits[0].get("label", title)) if hits else (None, title)


def wd_sparql(qid, session):
    """영화 QID의 촬영지 + 메타데이터를 한 번에 조회."""
    query = f"""
SELECT ?locLabel ?addr ?adminLabel ?countryLabel ?coord
  (SAMPLE(?yr) AS ?year)
  (GROUP_CONCAT(DISTINCT ?dl; separator=" | ") AS ?directors)
  (GROUP_CONCAT(DISTINCT ?cl; separator=" | ") AS ?cast)
WHERE {{
  VALUES ?film {{ wd:{qid} }}
  ?film wdt:P915 ?loc .
  OPTIONAL {{ ?film wdt:P577 ?date . BIND(YEAR(?date) AS ?yr) }}
  OPTIONAL {{ ?film wdt:P57 ?d . ?d rdfs:label ?dl . FILTER(LANG(?dl)="en") }}
  OPTIONAL {{ ?film wdt:P161 ?c . ?c rdfs:label ?cl . FILTER(LANG(?cl)="en") }}
  OPTIONAL {{ ?loc wdt:P6375 ?addr }}
  OPTIONAL {{ ?loc wdt:P131 ?admin . ?admin rdfs:label ?adminLabel . FILTER(LANG(?adminLabel)="en") }}
  OPTIONAL {{ ?loc wdt:P17 ?country . ?country rdfs:label ?countryLabel . FILTER(LANG(?countryLabel)="en") }}
  OPTIONAL {{ ?loc wdt:P625 ?coord }}
  ?loc rdfs:label ?locLabel . FILTER(LANG(?locLabel)="en")
}}
GROUP BY ?locLabel ?addr ?adminLabel ?countryLabel ?coord
""".strip()
    r = session.get(WD_SPARQL, params={"query": query, "format": "json"},
                    headers=WD_HEADERS, timeout=60)
    r.raise_for_status()
    return r.json()["results"]["bindings"]


def _coord_to_latlon(point):
    # "Point(lon lat)" → "lat,lon"
    m = re.match(r"Point\(([-0-9.]+)\s+([-0-9.]+)\)", point or "")
    if not m:
        return ""
    lon, lat = m.group(1), m.group(2)
    return f"{lat},{lon}"


def sparql_to_locations(bindings):
    """SPARQL bindings → 공통 스키마 촬영지 리스트 (순수 함수, 단위테스트 대상)."""
    locs = []
    for b in bindings:
        def g(k):
            return b.get(k, {}).get("value", "")
        area = ", ".join([x for x in (g("adminLabel"), g("countryLabel")) if x])
        locs.append({
            "fictional_name": "",                      # Wikidata P915 = 실제 장소(가상명 없음)
            "real_name": g("locLabel"),
            "scene_role": "",                          # Wikidata엔 장면 역할 없음
            "address": g("addr"),
            "area": area,
            "status": "",
            "coordinates": _coord_to_latlon(g("coord")),
            "sources": ["wikidata"],
        })
    return locs


def wikidata_extract(title, session):
    qid, label = wd_resolve_qid(title, session)
    if not qid:
        print(f"[warn] Wikidata에서 영화 못 찾음: {title}", file=sys.stderr)
        return None
    bindings = wd_sparql(qid, session)
    if not bindings:
        print(f"[info] P915 촬영지 없음: {title} ({qid})", file=sys.stderr)
        return None
    first = bindings[0]
    directors = [x for x in first.get("directors", {}).get("value", "").split(" | ") if x]
    cast = [x for x in first.get("cast", {}).get("value", "").split(" | ") if x]
    year = first.get("year", {}).get("value")
    return {
        "title": label,
        "year": int(year) if year and year.isdigit() else None,
        "directors": directors,
        "cast": cast,
        "regions": [],
        "source_url": f"https://www.wikidata.org/wiki/{qid}",
        "license": "CC0",
        "sources": ["wikidata"],
        "locations": sparql_to_locations(bindings),
    }


# ============================================================================
# 병합 (union + real_name 중복제거, 출처 기록)
# ============================================================================
def merge_sources(movieloc_movies, wikidata_movies):
    """제목 기준 병합. movieloc의 풍부한 설명 유지 + Wikidata 좌표/주소로 보강."""
    by_key = {}
    order = []

    def add(mv):
        key = _norm(mv["title"])
        if key not in by_key:
            by_key[key] = json.loads(json.dumps(mv))   # 복사
            order.append(key)
        else:
            tgt = by_key[key]
            tgt["sources"] = sorted(set(tgt.get("sources", [])) | set(mv.get("sources", [])))
            # 메타 빈칸 보강
            for f in ("year",):
                if not tgt.get(f) and mv.get(f):
                    tgt[f] = mv[f]
            for f in ("directors", "cast", "regions"):
                if not tgt.get(f) and mv.get(f):
                    tgt[f] = mv[f]
            # 촬영지 병합
            idx = {_norm(l["real_name"]): l for l in tgt["locations"]}
            for loc in mv["locations"]:
                k = _norm(loc["real_name"])
                if k and k in idx:
                    cur = idx[k]
                    cur["sources"] = sorted(set(cur.get("sources", [])) | set(loc.get("sources", [])))
                    # 빈칸만 보강 (movieloc 설명 우선)
                    for f in ("address", "area", "coordinates", "status",
                              "fictional_name", "scene_role"):
                        if not cur.get(f) and loc.get(f):
                            cur[f] = loc[f]
                else:
                    tgt["locations"].append(loc)
                    if k:
                        idx[k] = loc

    for mv in movieloc_movies:    # movieloc 먼저(설명 우선권)
        add(mv)
    for mv in wikidata_movies:
        add(mv)
    return [by_key[k] for k in order]


# ============================================================================
# 출력
# ============================================================================
def write_outputs(movies, out_base):
    payload = {
        "sources": {"movieloc": "movie-locations.com (rephrased facts)",
                    "wikidata": "Wikidata P915 (CC0)"},
        "extraction_policy": "객관적 사실만. scene_role 등 설명 100% 리프레이즈. "
                             "fictional/real 분리. 출처(sources)·라이선스 기록.",
        "movie_count": len(movies),
        "movies": movies,
    }
    with open(f"{out_base}.json", "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print(f"[ok] JSON 저장: {out_base}.json")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("[warn] openpyxl 미설치 → Excel 생략", file=sys.stderr)
        return

    wb = Workbook()
    hf = Font(name="Arial", bold=True, color="FFFFFF")
    hfill = PatternFill("solid", start_color="1F4E78")
    wrap = Alignment(vertical="top", wrap_text=True)

    s1 = wb.active; s1.title = "Movies"
    c1 = ["Title", "Year", "Directors", "Cast", "Regions", "Location count", "Sources", "Source URL"]
    s1.append(c1)
    for mv in movies:
        s1.append([mv.get("title"), mv.get("year"), ", ".join(mv.get("directors", [])),
                   ", ".join(mv.get("cast", [])), " | ".join(mv.get("regions", [])),
                   len(mv.get("locations", [])), ", ".join(mv.get("sources", [])),
                   mv.get("source_url")])
    for i, w in enumerate([24, 7, 20, 42, 28, 13, 16, 40], 1):
        s1.column_dimensions[chr(64 + i)].width = w

    s2 = wb.create_sheet("Locations")
    c2 = ["Title", "Year", "Fictional name (in film)", "Real place", "Scene / role",
          "Address", "Area", "Status", "Coordinates", "Sources"]
    s2.append(c2)
    for mv in movies:
        for loc in mv.get("locations", []):
            s2.append([mv.get("title"), mv.get("year"), loc.get("fictional_name", ""),
                       loc.get("real_name", ""), loc.get("scene_role", ""),
                       loc.get("address", ""), loc.get("area", ""), loc.get("status", ""),
                       loc.get("coordinates", ""), ", ".join(loc.get("sources", []))])
    for i, w in enumerate([18, 7, 24, 30, 38, 26, 24, 14, 18, 16], 1):
        s2.column_dimensions[chr(64 + i)].width = w

    for s, n in ((s1, len(c1)), (s2, len(c2))):
        for c in range(1, n + 1):
            cell = s.cell(1, c); cell.font = hf; cell.fill = hfill
        for row in s.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = wrap
        s.freeze_panes = "A2"
        s.auto_filter.ref = f"A1:{chr(64 + n)}{s.max_row}"

    wb.save(f"{out_base}.xlsx")
    print(f"[ok] Excel 저장: {out_base}.xlsx")


DEFAULT_URLS = [
    f"{BASE}/movies/h/Heat.php", f"{BASE}/movies/g/Godfather.php",
    f"{BASE}/movies/j/Jaws.php", f"{BASE}/movies/g/Gladiator.php",
    f"{BASE}/movies/m/Matrix.php", f"{BASE}/movies/v/Vertigo.php",
    f"{BASE}/movies/f/Fight-Club.php", f"{BASE}/movies/d/Die-Hard.php",
    f"{BASE}/movies/t/Taxi-Driver.php", f"{BASE}/movies/p/Pretty-Woman.php",
]


def run_movieloc(args):
    fetcher = PoliteFetcher(delay=args.delay)
    api_key = os.getenv("LLM_API_KEY")
    if not api_key:
        sys.exit("[error] movieloc 소스에는 환경변수 LLM_API_KEY가 필요합니다.")
    if args.urls:
        urls = args.urls
    elif args.letter:
        urls = movieloc_discover(fetcher, args.letter.lower(), args.limit)
        print(f"[info] '{args.letter}' 인덱스에서 {len(urls)}개 발견")
    else:
        urls = DEFAULT_URLS
    out = []
    for i, url in enumerate(urls, 1):
        print(f"[movieloc {i}/{len(urls)}] {url}")
        html = fetcher.get(url)
        if not html:
            continue
        try:
            out.append(llm_extract(clean_page_text(html), url, args.model, api_key, args.base_url))
        except Exception as e:
            print(f"[error] LLM 추출 실패: {url} - {e}", file=sys.stderr)
    return out


def run_wikidata(args):
    if not args.titles:
        sys.exit("[error] wikidata 소스에는 --titles \"제목\" ... 가 필요합니다.")
    session = requests.Session()
    out = []
    for i, title in enumerate(args.titles, 1):
        print(f"[wikidata {i}/{len(args.titles)}] {title}")
        try:
            mv = wikidata_extract(title, session)
            if mv:
                out.append(mv)
        except Exception as e:
            print(f"[error] Wikidata 조회 실패: {title} - {e}", file=sys.stderr)
        time.sleep(args.delay)
    return out


def main():
    ap = argparse.ArgumentParser(description="영화 촬영지 사실 수집 (movieloc + wikidata)")
    ap.add_argument("--source", choices=["movieloc", "wikidata", "both"], default="wikidata",
                    help="데이터 소스 (기본: wikidata, CC0라 가장 자유로움)")
    ap.add_argument("--titles", nargs="*", help="wikidata용 영화 제목 목록")
    ap.add_argument("--urls", nargs="*", help="movieloc용 페이지 URL 목록")
    ap.add_argument("--letter", help="movieloc A-Z 인덱스 자동 수집")
    ap.add_argument("--limit", type=int, default=20)
    ap.add_argument("--delay", type=float, default=2.0)
    ap.add_argument("--out", default="movie_locations")
    ap.add_argument("--model", default=os.getenv("LLM_MODEL", "gpt-4o-mini"))
    ap.add_argument("--base-url", default=os.getenv("LLM_BASE_URL", "https://api.openai.com/v1"))
    args = ap.parse_args()

    ml, wd = [], []
    if args.source in ("movieloc", "both"):
        ml = run_movieloc(args)
    if args.source in ("wikidata", "both"):
        wd = run_wikidata(args)

    movies = merge_sources(ml, wd) if args.source == "both" else (ml or wd)
    print(f"[done] {len(movies)}편, 촬영지 {sum(len(m['locations']) for m in movies)}개")
    write_outputs(movies, args.out)


if __name__ == "__main__":
    main()
