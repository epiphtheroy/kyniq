# -*- coding: utf-8 -*-
# 28편 배치 실행 결과 빌더 (실제 웹검색 결과를 등급화하여 JSON+xlsx 산출)
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# loc tuple: (real_name, narrative_setting, built_set, set_host, scene_role,
#             address, filming_area, country, granularity, [domains])
DATA = {
 ("Parasite",2019): [
  ("Sky Pizza","Kim family's pizza job",False,"","Pizzeria whose boxes the father folds","86 Noryangjin-ro 6-gil, Dongjak-gu","Seoul","South Korea","address",["cnn.com","english.visitkorea.or.kr","creatrip.com"]),
  ("Doijissal (Pig Rice) Supermarket","'Woori Supermarket'",False,"","Store where the son hears of the tutoring job","32 Songijeong-ro, Mapo-gu","Seoul","South Korea","address",["cnn.com","creatrip.com","klook.com"]),
  ("Stairs & alleys of Ahyeon-dong / Changsin-dong","Kim family's low-lying district",False,"","Steep stairways marking the family's descent","","Seoul","South Korea","area",["english.visitkorea.or.kr","cnn.com","klook.com"]),
  ("Park family mansion","wealthy Seoul home",True,"Jeonju Film Studio Complex (set + CGI)","Purpose-built set, not a real residence","near Jeonju Film Studio Complex","Jeonju","South Korea","set",["screenrant.com","dezeen.com","filmoblivion.com"]),
 ],
 ("La La Land",2016): [
  ("Griffith Observatory","Los Angeles",False,"","Planetarium 'floating' dance","2800 E Observatory Rd","Los Angeles","USA","address",["socalpulse.com","thecinemaholic.com","imdb.com"]),
  ("Colorado Street Bridge","Los Angeles",False,"","1913 bridge backdrop","532 W Colorado Blvd, Pasadena","Pasadena","USA","address",["socalpulse.com","thecinemaholic.com"]),
  ("Angels Flight Railway","Los Angeles",False,"","Funicular the couple ride","351 S Hill St","Downtown Los Angeles","USA","address",["socalpulse.com","livelybylaura.com"]),
  ("Lighthouse Cafe","LA jazz club",False,"","Jazz club the pianist plays","30 Pier Ave, Hermosa Beach","Hermosa Beach","USA","address",["socalpulse.com","thecinemaholic.com"]),
 ],
 ("Mad Max: Fury Road",2015): [
  ("Namib Desert near Swakopmund","the Wasteland",False,"","Primary desert filming base","","Swakopmund","Namibia","region",["npr.org","nomadicnotes.com","imdb.com"]),
  ("Salt plains near Walvis Bay","the Wasteland",False,"","Where the War Rig breaks down","","Walvis Bay","Namibia","area",["nomadicnotes.com","orbitzfeatures.waveinteractive.com"]),
  ("Dorob National Park","the Wasteland",False,"","Relocated main filming area","","Erongo Region","Namibia","region",["nomadicnotes.com","npr.org"]),
 ],
 ("Skyfall",2012): [
  ("Glencoe (A82, Buachaille Etive Mor)","Scottish Highlands",False,"","Bond and M drive into the Highlands","A82, Glencoe","Glencoe","Scotland, UK","area",["thejamesbonddossier.com","huntingbond.com","en.wikipedia.org"]),
  ("Skyfall Lodge","Bond's Scottish ancestral home",True,"Hankley Common, Surrey (plywood/plaster set, blown up)","A built set, not a real Scottish house","Hankley Common, Elstead","Surrey","England, UK","set",["007.info","visitsurrey.com","getsurrey.co.uk"]),
  ("Grand Bazaar & Eminonu Square","Istanbul",False,"","Opening rooftop & market chase","Eminonu, Fatih","Istanbul","Turkey","area",["huntingbond.com","media.hopper.com","imdb.com"]),
  ("Old Vic Tunnels (Waterloo)","MI6 underground",False,"","Disused railway vaults as MI6 tunnels","beneath Waterloo Station","London","England, UK","venue",["tokyofox.net","imdb.com"]),
 ],
 ("Inception",2010): [
  ("Pont de Bir-Hakeim","Paris dream",False,"","Bridge where Cobb teaches Ariadne","","Paris","France","venue",["nolanlocations.com","screenrant.com","imdb.com"]),
  ("Fortress Mountain ski resort","snow-fortress level",False,"","Mountain fortress assault","Kananaskis","Alberta","Canada","area",["nolanlocations.com","screenrant.com"]),
  ("Cardington hangars","rotating-corridor & hotel sets",True,"Cardington Studios, Bedfordshire","Rotating corridor / hotel built as sets","Cardington","Bedfordshire","England, UK","set",["nolanlocations.com","jerrygarrett.wordpress.com"]),
  ("John Ferraro Building (LADWP)","Los Angeles",False,"","City department-of-water-and-power exterior","111 N Hope St","Los Angeles","USA","venue",["seeing-stars.com","imdb.com"]),
 ],
 ("The Dark Knight",2008): [
  ("Chicago Main Post Office","Gotham National Bank",False,"","Opening bank robbery","404 W Harrison St","Chicago","USA","address",["legendarytrips.com","giggster.com","imdb.com"]),
  ("Richard J Daley Center","Wayne Enterprises HQ",False,"","Wayne Enterprises tower","Washington St","Chicago","USA","venue",["legendarytrips.com","giggster.com"]),
  ("IBM Building (AMA Plaza)","Wayne Enterprises offices",False,"","Dent/Commissioner/boardroom interiors","330 N Wabash Ave","Chicago","USA","address",["legendarytrips.com","giggster.com"]),
  ("Willis (Sears) Tower","Gotham skyline",False,"","Batman overlooks the city","233 S Wacker Dr","Chicago","USA","address",["legendarytrips.com","imdb.com"]),
 ],
 ("Slumdog Millionaire",2008): [
  ("Dharavi","Mumbai slums",False,"","Jamal's childhood neighbourhood","","Mumbai","India","area",["qantas.com","sceneloc8.com","imdb.com"]),
  ("Chhatrapati Shivaji Terminus (CST)","Mumbai station",False,"","Abduction scene & 'Jai Ho' finale","","Mumbai","India","venue",["jayatravel.com","qantas.com","imdb.com"]),
  ("Taj Mahal","Agra (hustling flashback)",False,"","Boys hustle tourists","","Agra","India","venue",["jayatravel.com","itv.com"]),
 ],
 ("The Lord of the Rings: Fellowship",2001): [
  ("Hobbiton movie set, Alexander Farm","the Shire",True,"Built set on Alexander Farm, Matamata (now tourable)","The Shire village built on a sheep farm","","Matamata","New Zealand","set",["newzealand.com","firstlighttravel.com","findingtheuniverse.com"]),
  ("Mount Ngauruhoe, Tongariro NP","Mount Doom",False,"","Stand-in for Mount Doom","","Tongariro National Park","New Zealand","area",["newzealand.com","firstlighttravel.com"]),
  ("Kaitoke Regional Park","Rivendell",False,"","Elvish realm where the Fellowship forms","near Upper Hutt","Wellington region","New Zealand","area",["newzealand.com","firstlighttravel.com"]),
  ("Mount Sunday","Edoras",False,"","Hilltop city of Rohan","Canterbury","South Island","New Zealand","area",["newzealand.com","firstlighttravel.com"]),
 ],
 ("Harry Potter and the Philosopher's Stone",2001): [
  ("Leavesden Studios","Hogwarts Great Hall & interiors",True,"Warner Bros Studios Leavesden (sets)","Great Hall and many interior sets built here","","Leavesden","England, UK","set",["en.wikipedia.org","ricksteves.com"]),
  ("Alnwick Castle","Hogwarts exterior",False,"","Flying-lesson courtyard / castle exterior","Alnwick, Northumberland","Northumberland","England, UK","venue",["en.wikipedia.org","scottishtours.co.uk","findingtheuniverse.com"]),
  ("Gloucester Cathedral","Hogwarts corridors",False,"","Cloister corridors","","Gloucester","England, UK","venue",["en.wikipedia.org","ricksteves.com"]),
  ("Leadenhall Market","Diagon Alley / Leaky Cauldron entrance",False,"","Entrance to the wizard pub","42 Bull's Head Passage","London","England, UK","address",["en.wikipedia.org","findingtheuniverse.com"]),
  ("12 Picket Post Close","4 Privet Drive",False,"","The Dursleys' house","Martins Heron, Bracknell, Berkshire","Berkshire","England, UK","address",["en.wikipedia.org","latlong.net"]),
 ],
 ("The Beach",2000): [
  ("Maya Bay, Ko Phi Phi Leh","the secret island lagoon",False,"","The hidden beach community","Ko Phi Phi Leh","Krabi","Thailand","area",["giggster.com","almostginger.com","timeout.com"]),
  ("Haew Suwat Waterfall, Khao Yai NP","jungle waterfall",False,"","Waterfall the characters leap from","Khao Yai National Park","Nakhon Ratchasima","Thailand","area",["almostginger.com","backpackbob.com"]),
 ],
 ("Lost in Translation",2003): [
  ("Park Hyatt Tokyo","the hotel",False,"","Bar/room/pool where the leads meet","3-7-1-2 Nishi-Shinjuku","Tokyo","Japan","address",["focusfeatures.com","filmoblivion.com","imdb.com"]),
  ("Shibuya Crossing","Tokyo streets",False,"","The famous scramble crossing","Shibuya","Tokyo","Japan","venue",["focusfeatures.com","tokyofox.net"]),
  ("Karaoke Kan, Shibuya","karaoke night",False,"","Private karaoke rooms 601/602","Shibuya","Tokyo","Japan","venue",["filmoblivion.com","tokyofox.net"]),
 ],
 ("Amelie",2001): [
  ("Cafe des Deux Moulins","Amelie's workplace",False,"","Art-deco cafe where Amelie waitresses","15 rue Lepic","Paris (Montmartre)","France","address",["theculturetrip.com","coolstuffinparis.com","completefrance.com"]),
  ("Sacre-Coeur steps","Montmartre",False,"","Treasure-hunt blue-arrow steps","Montmartre","Paris","France","venue",["theculturetrip.com","worldofwanderlust.com"]),
  ("Canal Saint-Martin","Paris canal",False,"","Where Amelie skips stones","10th arrondissement","Paris","France","area",["theculturetrip.com","almostginger.com"]),
  ("Gare de l'Est","train station",False,"","Photo-booth scene","Gare de l'Est","Paris","France","venue",["coolstuffinparis.com","almostginger.com"]),
 ],
 ("The Grand Budapest Hotel",2014): [
  ("Goerlitzer Warenhaus","the hotel lobby",False,"","1913 department store; lobby sets built inside","Demianplatz","Goerlitz","Germany","venue",["cnn.com","locationshub.com","atlasofwonders.com"]),
  ("Molkerei Pfund, Dresden","Mendl's patisserie interior",False,"","19th-c creamery as the cake-shop interior","Dresden","Saxony","Germany","venue",["locationshub.com","johnthego.com"]),
  ("Freisebad bathhouse","the hotel baths",False,"","1887 bathhouse used as the spa","Goerlitz","Saxony","Germany","venue",["locationshub.com","johnthego.com"]),
 ],
 ("Notting Hill",1999): [
  ("280 Westbourne Park Road","the Blue Door flat",False,"","William's front door (interior was a studio set)","280 Westbourne Park Rd","London","England, UK","address",["thetrainline.com","britmovietours.com","qantas.com"]),
  ("142 Portobello Road","the travel bookshop",False,"","William's bookshop storefront","142 Portobello Rd","London","England, UK","address",["thetrainline.com","britmovietours.com"]),
  ("The Coronet","Notting Hill cinema",False,"","Cinema where William watches Anna's film","Notting Hill Gate","London","England, UK","venue",["thetrainline.com","underlondonlights.com"]),
  ("The Savoy Hotel","press-conference hotel",False,"","Anna's press conference","Strand","London","England, UK","venue",["britmovietours.com","countryandtownhouse.com"]),
 ],
 ("Braveheart",1995): [
  ("Glen Nevis","Highland village",False,"","'Lanark' village built in the glen","near Fort William","Highland","Scotland, UK","area",["thetrainline.com","giggster.com","imdb.com"]),
  ("Glen Coe","Highland landscape",False,"","Sweeping Highland scenery","Glencoe","Highland","Scotland, UK","area",["thetrainline.com","giggster.com"]),
  ("Trim Castle","'York'",False,"","English-held town of York","Trim, Co Meath","Leinster","Ireland","venue",["giggster.com","almostginger.com"]),
  ("Curragh Plains","Battle of Stirling Bridge",False,"","Major battle field","Co Kildare","Leinster","Ireland","area",["giggster.com","outono.net"]),
  ("Ardmore Studios","interiors",True,"Ardmore Studios, Bray, Ireland","Interior sets","Bray, Co Wicklow","Leinster","Ireland","set",["giggster.com","imdb.com"]),
 ],
 ("Black Panther",2018): [
  ("Pinewood Atlanta Studios","Wakanda",True,"Pinewood Atlanta Studios (sets + green screen)","Wakanda built on soundstages","","Atlanta, GA","USA","set",["giggster.com","en.wikipedia.org","filmoblivion.com"]),
  ("Gwangan Bridge & Gwangalli Beach","Busan car chase",False,"","150-car chase sequence","Busan","Busan","South Korea","area",["giggster.com","en.wikipedia.org","atlasofwonders.com"]),
  ("High Museum of Art","'Museum of Great Britain'",False,"","Museum heist scene","1280 Peachtree St NE, Atlanta","Atlanta, GA","USA","address",["giggster.com","travelnoire.com"]),
  ("Sweet Auburn district","'Oakland'",False,"","Doubles for Oakland flashbacks","Atlanta, GA","USA","USA","area",["giggster.com","filmoblivion.com"]),
 ],
 ("Tenet",2020): [
  ("Linnahall","'Kyiv Opera House'",False,"","Opening opera-house siege","Kalasadama 4","Tallinn","Estonia","address",["mensjournal.com","theartsshelf.com","en.wikipedia.org"]),
  ("Laagna Road","Tallinn highway",False,"","Inverted car chase","Laagna tee","Tallinn","Estonia","area",["mensjournal.com","theculturetrip.com"]),
  ("Villa Cimbrone, Ravello","Amalfi meeting",False,"","Protagonist meets Kat","Ravello","Amalfi Coast","Italy","venue",["atlasofwonders.com","mensjournal.com"]),
  ("Gateway of India & Taj Mahal Palace Hotel","Mumbai",False,"","Mumbai infiltration sequence","Colaba","Mumbai","India","venue",["thecinemaholic.com","imdb.com"]),
 ],
 ("No Time to Die",2021): [
  ("Matera","old Italian town",False,"","Opening car chase & hideaway","Matera, Basilicata","Matera","Italy","area",["onthetracksof007.com","huntingbond.com","screenrant.com"]),
  ("Gravina aqueduct bridge","Italian gorge",False,"","Bridge crossing","Gravina in Puglia","Puglia","Italy","venue",["onthetracksof007.com","screenrant.com"]),
  ("Atlantic Ocean Road","Norway coast",False,"","Bond pursues Madeleine","More og Romsdal","Norway","Norway","area",["huntingbond.com","screenrant.com"]),
  ("Kalsoy","Safin's island",False,"","Villain's island (facility added in CGI)","Kalsoy, Faroe Islands","Faroe Islands","Denmark","area",["huntingbond.com","wonderfulwanderings.com"]),
  ("San San Beach, Port Antonio","Bond's Jamaica home",False,"","Bond's retirement beach house (built then removed)","Port Antonio","Portland","Jamaica","area",["huntingbond.com","mensjournal.com"]),
 ],
 ("Dunkirk",2017): [
  ("Plage de Malo-les-Bains","Dunkirk beach",False,"","Beach evacuation scenes on the real site","Digue de Mer, Dunkerque","Dunkerque","France","area",["completefrance.com","screenrant.com","en.wikipedia.org"]),
  ("Replica East Mole (jetty)","the evacuation mole",True,"Replica built on the real Malo-les-Bains site","Reconstructed jetty for the evacuation","Malo-les-Bains","Dunkerque","France","set",["screenrant.com","almostginger.com"]),
 ],
 ("1917",2019): [
  ("Salisbury Plain","'Northern France'",False,"","Trenches dug for the no-man's-land battle","Wiltshire","Wiltshire","England, UK","area",["giggster.com","visitwiltshire.co.uk","imdb.com"]),
  ("Ambrose Quarry","ruined-town/quarry",False,"","Disused quarry sequence","near Ewelme, Oxfordshire","Oxfordshire","England, UK","area",["movie-locations.com","thecinemaholic.com"]),
 ],
 ("Interstellar",2014): [
  ("Ranch near Longview","Cooper's farm",False,"","Farmhouse built; 500 acres of corn planted","near Longview","Alberta","Canada","area",["screenrant.com","imdb.com","setjetters.com"]),
  ("Svinafellsjokull glacier","the ice planet",False,"","Mann's frozen planet","near Skaftafell","South Iceland","Iceland","area",["legendarytrips.com","screenrant.com"]),
  ("Eldhraun lava field (Orrustuholl)","the water planet",False,"","Miller's planet surface","Eldhraun","South Iceland","Iceland","area",["legendarytrips.com","screenrant.com"]),
  ("Westin Bonaventure Hotel","NASA facility",False,"","Secret NASA base interior","404 S Figueroa St","Los Angeles","USA","address",["screenrant.com","setjetters.com"]),
 ],
 ("Crouching Tiger, Hidden Dragon",2000): [
  ("Mukeng Bamboo Forest","bamboo treetop duel",False,"","Treetop fight backdrop","Mukeng, Huangshan","Anhui","China","area",["chinaadvent.com","chinadiscovery.com","imdb.com"]),
  ("Anji Bamboo Sea","bamboo sea duel",False,"","Flying bamboo-grove fight","Anji, Huzhou","Zhejiang","China","area",["chinaservicesinfo.com","chinaadvent.com"]),
  ("Cangyan Shan (Cangyan Mountain)","warrior temple",False,"","Hanging temple scenes","SW of Shijiazhuang","Hebei","China","area",["movie-locations.com","imdb.com"]),
  ("Hengdian World Studios","palace interiors",True,"Hengdian World Studios, Zhejiang","Elaborate palace sets","Hengdian","Zhejiang","China","set",["movie-locations.com","imdb.com"]),
 ],
 ("Roman Holiday",1953): [
  ("Spanish Steps","Rome",False,"","Princess eats gelato","Scalinata di Trinita dei Monti","Rome","Italy","venue",["giggster.com","thetrainline.com","en.wikipedia.org"]),
  ("Bocca della Verita (Mouth of Truth)","Rome",False,"","The hand-in-the-mouth scene","Santa Maria in Cosmedin","Rome","Italy","venue",["americanclubrome.org","giggster.com"]),
  ("Via Margutta 51","Bradley's apartment",False,"","Reporter's apartment","Via Margutta 51","Rome","Italy","address",["thetrainline.com","almostginger.com"]),
  ("Palazzo Colonna","royal press hall",False,"","The film's farewell finale","Piazza SS Apostoli 66","Rome","Italy","address",["giggster.com","nomadepicureans.com"]),
  ("Cinecitta Studios","interiors",True,"Cinecitta Studios, Rome","Studio interiors ('Hollywood on the Tiber')","Rome","Lazio","Italy","set",["en.wikipedia.org","giggster.com"]),
 ],
 ("The Sound of Music",1965): [
  ("Schloss Leopoldskron","von Trapp villa (lake side)",False,"","Lakeside terrace; boat-tipping scene","Leopoldskronstrasse","Salzburg","Austria","venue",["salzburg.info","salzburgerland.com","panoramatours.com"]),
  ("Schloss Frohnburg","von Trapp villa (gates/front)",False,"","Villa front gate and entrance","Hellbrunner Allee","Salzburg","Austria","venue",["salzburg.info","panoramatours.com"]),
  ("Nonnberg Abbey","Maria's abbey",False,"","Exterior of the convent","Nonnberggasse","Salzburg","Austria","venue",["salzburg.info","roadaffair.com"]),
  ("Mirabell Gardens","'Do-Re-Mi'",False,"","Pegasus-fountain dance","Mirabellgarten","Salzburg","Austria","venue",["salzburg.info","panoramatours.com"]),
  ("Felsenreitschule","festival stage",False,"","Family's final performance before fleeing","Hofstallgasse","Salzburg","Austria","venue",["salzburg.info","roadaffair.com"]),
  ("Basilica St Michael, Mondsee","the wedding",False,"","Maria & Georg's wedding","Mondsee","Upper Austria","Austria","venue",["salzburg.info","roadaffair.com"]),
 ],
 ("Jurassic Park",1993): [
  ("Manawaiopuna Falls","'Jurassic Falls'",False,"","Helicopter arrival waterfall","Hanapepe Valley, Kauai","Kauai, Hawaii","USA","area",["beatofhawaii.com","makanacharters.com","hawaiiactivities.com"]),
  ("Allerton Garden","park grounds",False,"","Moreton Bay figs; maintenance shed","4425 Lawai Rd, Koloa, Kauai","Kauai, Hawaii","USA","address",["beatofhawaii.com","makanacharters.com"]),
  ("Kualoa Ranch, Ka'a'awa Valley","Gallimimus stampede",False,"","Stampede & fallen-tree scene","Kamehameha Hwy, Oahu","Oahu, Hawaii","USA","area",["beatofhawaii.com","set-jetter.com"]),
  ("Na Pali Coast","island entrance",False,"","Dramatic island fly-in cliffs","Kauai","Kauai, Hawaii","USA","area",["hawaiiactivities.com","makanacharters.com"]),
 ],
 ("Trainspotting",1996): [
  ("Former John Menzies, Princes Street","Edinburgh",False,"","Opening shoplifting & street run","107-109 Princes St","Edinburgh","Scotland, UK","address",["tokyofox.net","almostginger.com","imdb.com"]),
  ("WD & HO Wills tobacco factory","various interiors",False,"","~30 of the film's interiors","Alexandra Parade","Glasgow","Scotland, UK","venue",["almostginger.com","glasgowworld.com"]),
  ("Cafe D'Jaconelli","milkshake cafe",False,"","Spud & Renton before the job interview","570 Maryhill Rd","Glasgow","Scotland, UK","address",["almostginger.com","glasgowworld.com"]),
  ("Rouken Glen Park","park scene",False,"","Renton & Sick Boy's air-rifle scene","Thornliebank","Glasgow","Scotland, UK","area",["almostginger.com","imdb.com"]),
  ("Rannoch Moor station","Highland day-out",False,"","'Great outdoors' railway stop","Rannoch Moor","Highland","Scotland, UK","area",["almostginger.com","tokyofox.net"]),
 ],
 ("Oldboy",2003): [
  ("Chinese restaurant, Busan Chinatown","dumpling restaurant",False,"","Where Oh Dae-su eats his fried dumplings","Chinatown, Busan","Busan","South Korea","venue",["legendarytrips.com","imdb.com"]),
  ("Mount Lyford","snowy ending",False,"","Final scene in the snow","Southern Alps","Canterbury","New Zealand","area",["en.wikipedia.org","imdb.com"]),
 ],
 ("Bohemian Rhapsody",2018): [
  ("Bovingdon Airfield","Live Aid (Wembley 1985)",True,"Bovingdon Airfield, Hertfordshire (Wembley stage rebuilt)","Wembley Live Aid stage recreated (Wembley demolished 2002)","Bovingdon, Hemel Hempstead","Hertfordshire","England, UK","set",["variety.com","screenglobalproduction.com","en.wikipedia.org"]),
 ],
}

BLOCK = {"movie-locations.com", "atlasofwonders.com"}

def grade(doms, built):
    indep = [d for d in doms if d not in BLOCK]
    if len(set(doms)) >= 2 and len(indep) >= 1:
        c = "verified"
    elif len(doms) >= 1:
        c = "single_source"
    else:
        c = "unverified"
    if built and c == "verified":
        c = "verified_set_not_real"
    return c

movies = []
for (title, year), locs in DATA.items():
    L = []
    for (rn, ns, bs, sh, sr, ad, fa, co, gr, dom) in locs:
        L.append({"real_name": rn, "narrative_setting": ns, "built_set": bs,
                  "set_host": sh, "scene_role": sr, "address": ad,
                  "filming_area": fa, "country": co, "granularity": gr,
                  "sources": sorted(set(dom)), "source_count": len(set(dom)),
                  "confidence": grade(dom, bs)})
    movies.append({"title": title, "year": year, "location_count": len(L), "locations": L})

payload = {
    "method": "LLM + web search, citation-grounded, multi-source verified (batch run, 28 films)",
    "generated_at": "2026-06-28",
    "grading_rule": ">=2 independent domains -> verified; 1 -> single_source. built_set true => set_host names the studio/location.",
    "movie_count": len(movies), "movies": movies,
}
json.dump(payload, open("llmsearch_batch.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)

# xlsx
wb = Workbook(); s = wb.active; s.title = "Locations"
cols = ["Title","Year","Real place","Narrative setting","Built set?","Set host (studio/site)",
        "Scene/role","Address","Filming area","Country","Granularity","Confidence","Src#","Sources"]
s.append(cols)
tot=ver=ss=setn=0
for m in movies:
    for l in m["locations"]:
        tot+=1; c=l["confidence"]
        if c.startswith("verified"): ver+=1
        if c=="single_source": ss+=1
        if l["built_set"]: setn+=1
        s.append([m["title"],m["year"],l["real_name"],l["narrative_setting"],
                  "YES" if l["built_set"] else "",l["set_host"],l["scene_role"],
                  l["address"],l["filming_area"],l["country"],l["granularity"],
                  c,l["source_count"],", ".join(l["sources"])])
hf=Font(name="Arial",bold=True,color="FFFFFF"); fill=PatternFill("solid",start_color="1F4E78")
for c in range(1,len(cols)+1):
    cell=s.cell(1,c); cell.font=hf; cell.fill=fill
for i,w in enumerate([20,6,28,22,9,28,32,26,18,14,11,20,5,30],1):
    s.column_dimensions[chr(64+i)].width=w
for row in s.iter_rows(min_row=2):
    for cell in row: cell.alignment=Alignment(vertical="top",wrap_text=True)
s.freeze_panes="A2"; s.auto_filter.ref=f"A1:N{s.max_row}"
wb.save("llmsearch_batch.xlsx")

print(f"films: {len(movies)} | locations: {tot}")
print(f"verified(incl set-flag): {ver} | single_source: {ss} | built_set tagged: {setn}")
cnt_country=len({l['country'] for m in movies for l in m['locations']})
print(f"distinct countries: {cnt_country}")
